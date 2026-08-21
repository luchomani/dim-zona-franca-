# -*- coding: utf-8 -*-
"""
Procesador Masivo de Declaraciones de Importación (DIM - Formulario 500 DIAN)
================================================================================
Panel Corporativo Optimizado — Zona Franca
"""

import io
import re
import zipfile
import os
import base64
from datetime import datetime

import fitz  # PyMuPDF
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configuración general y Estilos Corporativos con Fondo Personalizado
# --------------------------------------------------------------------------

st.set_page_config(
    page_title="Procesador DIM | Zona Franca",
    page_icon="🏢",
    layout="wide",
)

fondo_css = ""
if os.path.exists("Fondo ZFC.png"):
    with open("Fondo ZFC.png", "rb") as f:
        fondo_bytes = f.read()
    fondo_base64 = base64.b64encode(fondo_bytes).decode()
    fondo_css = f"""
    .stApp {{
        background-image: linear-gradient(rgba(244, 247, 246, 0.9), rgba(244, 247, 246, 0.9)), url("data:image/png;base64,{fondo_base64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
    }}
    """
else:
    fondo_css = """
    .stApp {
        background-color: #F4F7F6;
    }
    """

st.markdown(f"""
<style>
    :root {{
        --zf-green-dark: #1B4D3E;
        --zf-green-medium: #2C6B56;
        --zf-olive: #8A9A28;
        --zf-card-bg: #FFFFFF;
        --zf-text-main: #2C3E50;
    }}

    {fondo_css}

    h1, h2, h3 {{
        color: var(--zf-green-dark) !important;
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
    }}

    div[data-testid="stVerticalBlock"] > div[style*="border"] {{
        background-color: var(--zf-card-bg);
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
        border: 1px solid #E1E8E5 !important;
        padding: 20px;
    }}

    .stButton>button {{
        background-color: var(--zf-green-dark);
        color: white;
        border-radius: 6px;
        border: none;
        font-weight: 600;
        padding: 0.5rem 1rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }}

    .stButton>button:hover {{
        background-color: var(--zf-green-medium);
        color: white;
        border: none;
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }}

    div[data-testid="stMetricValue"] {{
        color: var(--zf-green-dark);
        font-weight: 700;
    }}
    div[data-testid="stMetricLabel"] {{
        color: #556B2F;
        font-weight: 600;
    }}

    .stDataFrame {{
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid #E1E8E5;
    }}
</style>
""", unsafe_allow_html=True)

# LISTA OFICIAL DE COLUMNAS
COLUMNAS = [
    "Número de formulario",
    "NIT Importador",
    "Razón Social Importador",
    "Factura",
    "Manifiesto de carga",
    "Documento de transporte",
    "Cod. País Procedencia",
    "Cod. Modo Transporte",
    "Código de Bandera",
    "Tasa de Cambio",
    "Subpartida Arancelaria",
    "Cod. País Origen",
    "Cod. País Compra",
    "Peso Bruto (Kgs)",
    "Peso Neto (Kgs)",
    "Código de Embalaje",
    "No. Bultos",
    "Valor FOB (USD)",
    "Sumatoria Fletes/Seguros/Otros (USD)",
    "Acta de Inspección No.",
    "Levante No.",
    "Fecha del Levante",
    "Archivo",
    "Campos_no_encontrados",
]

MONTO = r"[\d\.,]+"
ENTERO_MILES = r"[\d\.,]+"

# --------------------------------------------------------------------------
# Parser Numérico y Extracción
# --------------------------------------------------------------------------

def limpiar_monto(val_str):
    if not val_str:
        return 0.0
    val_str = str(val_str).strip()
    val_str = re.sub(r'[^\d.,]', '', val_str)
    if not val_str:
        return 0.0

    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    else:
        partes = val_str.split('.')
        if len(partes) == 2:
            if len(partes[1]) == 3 and len(partes[0]) <= 3:
                val_str = val_str.replace('.', '')
        elif len(partes) > 2:
            if len(partes[-1]) == 2:
                val_str = "".join(partes[:-1]) + "." + partes[-1]
            else:
                val_str = val_str.replace('.', '')

    try:
        return float(val_str)
    except ValueError:
        return 0.0

def _buscar(patron, texto, flags=re.IGNORECASE, grupo=1):
    m = re.search(patron, texto, flags)
    if m:
        try:
            return m.group(grupo).strip()
        except IndexError:
            return None
    return None

def dividir_dims(texto: str):
    partes = re.split(r"(?=Declaraci[oó]n de Importaci[oó]n)", texto, flags=re.IGNORECASE)
    return [p for p in partes if re.search(r"N[uú]mero de formulario", p, re.IGNORECASE)]

def extraer_campos_dim(chunk_texto: str, texto_completo: str, nombre_archivo: str) -> dict:
    faltantes = []

    def campo(nombre, patron, grupo=1, default=""):
        valor = _buscar(patron, chunk_texto, grupo=grupo)
        if not valor:
            valor = _buscar(patron, texto_completo, grupo=grupo)
        if not valor:
            faltantes.append(nombre)
            return default
        return " ".join(valor.split())

    numero_formulario = campo("Número de formulario", r"4\s*\.\s*N[uú]mero de formulario\s*\n?\s*(\S+)")
    nit_importador = campo("NIT Importador", r"5\s*\.\s*N[uú]mero de Identificaci[oó]n Tributaria \(NIT\)\s*(\d{9,10})")
    razon_social = campo("Razón Social Importador", r"11\s*\.\s*Apellidos y nombres o Raz[oó]n Social\s*([^\n]+)")
    factura = campo("Factura", r"51\s*\.\s*No\.\s*de\s*factura\s*\n\s*(\S+)")
    
    manifiesto_carga = campo("Manifiesto de carga", r"42\s*\.?\s*Manifiesto\s+de\s+carga\s*(?:No\.?\s*)?([A-Za-z0-9\-]+)")
    if not manifiesto_carga:
        manifiesto_carga = campo("Manifiesto de carga", r"42\s*\.?\s*Manifiesto\s+de\s+carga[^\n]*\n\s*(?:No\.?\s*)?([A-Za-z0-9\-]+)")

    documento_transporte = campo("Documento de transporte", r"44\s*\.?\s*Documento\s+de\s+transporte\s*(?:No\.?\s*)?([A-Za-z0-9\-]+)")
    if not documento_transporte:
        documento_transporte = campo("Documento de transporte", r"44\s*\.?\s*Documento\s+de\s+transporte[^\n]*\n\s*(?:No\.?\s*)?([A-Za-z0-9\-]+)")

    cod_pais_procedencia = campo("Cod. País Procedencia", r"53\s*\.\s*(?:C[oó]d\.?\s*)?pa[ií]s\s+(?:de\s+)?procedencia\s*([A-Za-z0-9]{2,3})")
    cod_modo_transporte = campo("Cod. Modo Transporte", r"54\s*\.\s*Cod\.\s*Modo\s*Transporte\s*(\d)")
    codigo_bandera = campo("Código de Bandera", r"55\s*\.\s*C[oó]digo\s+(?:de\s+)?bandera\s*([A-Za-z0-9]{2,3})")
    tasa_cambio = campo("Tasa de Cambio", r"Tasa de cambio\s*\$?\s*cvs\.?\s*\n?\s*([\d.,]+)")
    subpartida = campo("Subpartida Arancelaria", r"59\s*\.\s*Subpartida arancelaria\s*(\d{10})")
    cod_pais_origen = campo("Cod. País Origen", r"66\s*\.\s*(?:C[oó]d\.?\s*)?pa[ií]s\s+(?:de\s+)?origen\s*([A-Za-z0-9]{2,3})")
    cod_pais_compra = campo("Cod. País Compra", r"70\s*\.\s*Cod\s*\.\s*pa[ií]s\s*\n?\s*compra\s*(\d{2,3})")
    codigo_embalaje = campo("Código de Embalaje", r"73\s*\.\s*C[oó]digo\s*\n?\s*embalaje\s*([A-Za-z0-9]{1,4})")
    
    peso_bruto = limpiar_monto(campo("Peso Bruto (Kgs)", r"71\s*\.\s*Peso bruto kgs\.\s*dcms\.\s*(" + MONTO + ")"))
    peso_neto = limpiar_monto(campo("Peso Neto (Kgs)", r"72\s*\.\s*Peso neto kgs\.\s*dcms\.\s*(" + MONTO + ")"))
    valor_fob = limpiar_monto(campo("Valor FOB (USD)", r"78\s*\.\s*Valor FOB USD\s*(" + MONTO + ")"))
    sumatoria_fletes = limpiar_monto(campo("Sumatoria Fletes/Seguros/Otros (USD)", r"82\s*\.\s*Sumatoria de fletes,?\s*seguros\s*\n?\s*y otros gastos USD\s*(" + MONTO + ")"))
    
    n_bultos_str = campo("No. Bultos", r"74\s*\.\s*No\.\s*bultos\s*(" + ENTERO_MILES + ")")
    try:
        no_bultos = int(re.sub(r'[^\d]', '', n_bultos_str)) if n_bultos_str else 0
    except ValueError:
        no_bultos = 0

    acta_inspeccion = ""
    m_acta = re.search(r"ACTA\s+DE\s+INSPECCI[OÓ]N\s*(?:No\.?|Número)?\s*[:\.]?\s*([0-9]{8,15})", texto_completo, re.IGNORECASE)
    if m_acta:
        acta_inspeccion = m_acta.group(1).strip()

    levante_no = ""
    m_lev_box = re.search(r"134\.?\s*Levante\s+No\.?\s*([0-9]{8,15})", chunk_texto, re.IGNORECASE)
    if m_lev_box:
        levante_no = m_lev_box.group(1).strip()
    
    if not levante_no:
        m_lev_gen = re.search(r"(?:Levante|Auto(?:rización)?)\s*(?:No\.?|Número)?\s*[:\.]?\s*([0-9]{8,15})", texto_completo, re.IGNORECASE)
        if m_lev_gen:
            levante_no = m_lev_gen.group(1).strip()

    if not levante_no:
        faltantes.append("Levante No.")

    fecha_levante = campo("Fecha del Levante", r"135\.?\s*Fecha[^\d\n]*(\d{4}\s*[-/\.]\s*\d{2}\s*[-/\.]\s*\d{2})")
    if not fecha_levante:
        m_fec = re.search(r"\b(20\d{2}[-/\.](?:0[1-9]|1[0-2])[-/\.](?:0[1-9]|[12]\d|3[01]))\b", texto_completo)
        if m_fec:
            fecha_levante = m_fec.group(1)
            if "Fecha del Levante" in faltantes:
                faltantes.remove("Fecha del Levante")

    if fecha_levante:
        fecha_levante = re.sub(r"\s+", "", fecha_levante)
        fecha_levante = re.sub(r"[/.]", "-", fecha_levante)

    return {
        "Número de formulario": numero_formulario,
        "NIT Importador": nit_importador,
        "Razón Social Importador": razon_social,
        "Factura": factura,
        "Manifiesto de carga": manifiesto_carga,
        "Documento de transporte": documento_transporte,
        "Cod. País Procedencia": cod_pais_procedencia,
        "Cod. Modo Transporte": cod_modo_transporte,
        "Código de Bandera": codigo_bandera,
        "Tasa de Cambio": tasa_cambio,
        "Subpartida Arancelaria": subpartida,
        "Cod. País Origen": cod_pais_origen,
        "Cod. País Compra": cod_pais_compra,
        "Peso Bruto (Kgs)": peso_bruto,
        "Peso Neto (Kgs)": peso_neto,
        "Código de Embalaje": codigo_embalaje,
        "No. Bultos": no_bultos,
        "Valor FOB (USD)": valor_fob,
        "Sumatoria Fletes/Seguros/Otros (USD)": sumatoria_fletes,
        "Acta de Inspección No.": acta_inspeccion,
        "Levante No.": levante_no,
        "Fecha del Levante": fecha_levante,
        "Archivo": nombre_archivo,
        "Campos_no_encontrados": ", ".join(faltantes) if faltantes else "",
    }

def extraer_texto_pdf(data: bytes) -> str:
    with fitz.open(stream=data, filetype="pdf") as doc:
        return "\n".join(page.get_text() for page in doc)

def obtener_pdfs_desde_upload(uploaded_file):
    nombre = uploaded_file.name
    contenido = uploaded_file.read()

    if nombre.lower().endswith(".zip"):
        pdfs = []
        with zipfile.ZipFile(io.BytesIO(contenido)) as zf:
            for info in zf.infolist():
                if info.filename.lower().endswith(".pdf") and not info.is_dir():
                    pdfs.append((info.filename.split("/")[-1], zf.read(info.filename)))
        return pdfs
    elif nombre.lower().endswith(".pdf"):
        return [(nombre, contenido)]
    else:
        return []

def procesar_archivos(uploaded_files, progress_callback=None) -> pd.DataFrame:
    filas = []
    tareas = []

    for uf in uploaded_files:
        tareas.extend(obtener_pdfs_desde_upload(uf))

    total = max(len(tareas), 1)
    for i, (nombre_pdf, data) in enumerate(tareas, start=1):
        try:
            texto_completo = extraer_texto_pdf(data)
            dim_chunks = dividir_dims(texto_completo)
            if not dim_chunks:
                dim_chunks = [texto_completo]
            for chunk in dim_chunks:
                fila = extraer_campos_dim(chunk, texto_completo, nombre_pdf)
                filas.append(fila)
        except Exception as exc: 
            fila = {c: "" for c in COLUMNAS}
            fila["Archivo"] = nombre_pdf
            fila["Campos_no_encontrados"] = f"ERROR AL PROCESAR: {exc}"
            filas.append(fila)
        if progress_callback:
            progress_callback(i / total, nombre_pdf)

    if not filas:
        return pd.DataFrame(columns=COLUMNAS)

    df = pd.DataFrame(filas, columns=COLUMNAS)

    if "Número de formulario" in df.columns:
        df["Número de formulario"] = df["Número de formulario"].astype(str).str.strip()
        df = df[
            df["Número de formulario"].notna() & 
            (df["Número de formulario"] != "") & 
            (df["Número de formulario"].str.lower() != "nan")
        ]
        df = df.drop_duplicates(subset=["Número de formulario"], keep="last").reset_index(drop=True)

    if "Levante No." in df.columns:
        df["Levante No."] = df["Levante No."].astype(str).str.strip()

    return df

# --------------------------------------------------------------------------
# Exportación a Excel y CSV
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="D9E1D9", end_color="D9E1D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def generar_excel(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    df_export = df.copy()

    if not df_export.empty:
        total_row = {c: "" for c in df_export.columns}
        total_row["Número de formulario"] = "TOTALES CONSOLIDADOS"
        total_row["Valor FOB (USD)"] = df_export["Valor FOB (USD)"].sum()
        total_row["Sumatoria Fletes/Seguros/Otros (USD)"] = df_export["Sumatoria Fletes/Seguros/Otros (USD)"].sum()
        total_row["Peso Bruto (Kgs)"] = df_export["Peso Bruto (Kgs)"].sum()
        total_row["Peso Neto (Kgs)"] = df_export["Peso Neto (Kgs)"].sum()
        total_row["No. Bultos"] = df_export["No. Bultos"].sum()
        
        df_export = pd.concat([df_export, pd.DataFrame([total_row])], ignore_index=True)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="DIM")
        ws = writer.sheets["DIM"]

        n_filas = df_export.shape[0]
        n_cols = df_export.shape[1]

        for col_idx in range(1, n_cols + 1):
            celda = ws.cell(row=1, column=col_idx)
            celda.fill = HEADER_FILL
            celda.font = HEADER_FONT
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = THIN_BORDER

        for row_idx in range(2, n_filas + 2):
            for col_idx in range(1, n_cols + 1):
                celda = ws.cell(row=row_idx, column=col_idx)
                celda.border = THIN_BORDER
                celda.alignment = Alignment(vertical="top", wrap_text=True)
                
                if isinstance(celda.value, (int, float)):
                    celda.number_format = '#,##0.00'
                    
                if row_idx == n_filas + 1:
                    celda.fill = TOTAL_FILL
                    celda.font = Font(bold=True)

        for col_idx, columna in enumerate(df_export.columns, start=1):
            longitudes = [len(str(columna))] + [len(str(v)) for v in df_export[columna].astype(str).tolist()]
            ancho = min(max(longitudes) + 3, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_filas}"
        ws.freeze_panes = "A2"

    return buffer.getvalue()

def generar_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, sep=",", encoding="utf-8-sig").encode("utf-8-sig")

# --------------------------------------------------------------------------
# Interfaz de Usuario Corporativa (Streamlit)
# --------------------------------------------------------------------------

col_logo, col_titulo = st.columns([1.3, 3.7])

with col_logo:
    logo_encontrado = False
    for filename in ["LOGO ZFS-ZFC.jpeg", "logo.jpeg", "logo.jpg", "logo.png"]:
        if os.path.exists(filename):
            st.image(filename, width=240)
            logo_encontrado = True
            break
    
    if not logo_encontrado:
        st.info("💡 Sube tu imagen de logo al repositorio como `LOGO ZFS-ZFC.jpeg`.")

with col_titulo:
    st.markdown("### Módulo de Gestión Aduanera")
    st.markdown("**Procesador Masivo de Declaraciones de Importación — Formulario 500**")
    st.caption("Zona Franca de Cúcuta | Operada por Zona Franca Santander")

st.divider()

if "df_resultado_dim" not in st.session_state:
    st.session_state.df_resultado_dim = pd.DataFrame(columns=COLUMNAS)

with st.container():
    st.subheader("1. Carga de Documentación Aduanera")
    uploaded_files = st.file_uploader(
        "Arrastra y suelta tus archivos PDF individuales o paquetes comprimidos en formato .ZIP",
        type=["pdf", "zip"],
        accept_multiple_files=True,
        key="dim_uploader",
    )

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        procesar = st.button("🚀 Procesar Documentos", type="primary", use_container_width=True)
    with col_btn2:
        limpiar = st.button("🧹 Limpiar y Reiniciar Panel", use_container_width=True)

if limpiar:
    st.session_state.df_resultado_dim = pd.DataFrame(columns=COLUMNAS)
    if "dim_uploader" in st.session_state:
        del st.session_state["dim_uploader"]
    st.rerun()

if procesar:
    if not uploaded_files:
        st.warning("Por favor carga al menos un archivo PDF o ZIP antes de ejecutar el procesamiento.")
    else:
        progreso = st.progress(0.0, text="Iniciando motor de extracción...")

        def _cb(pct, nombre):
            progreso.progress(pct, text=f"Procesando archivo: {nombre}")

        with st.spinner("Analizando y extrayendo campos clave de las DIMs..."):
            df = procesar_archivos(uploaded_files, progress_callback=_cb)

        progreso.empty()
        st.session_state.df_resultado_dim = df
        st.success(f"✅ Extracción exitosa. Se han consolidado {len(df)} declaración(es) única(s).")

# --------------------------------------------------------------------------
# Visualización de Resultados y Métricas Corporativas
# --------------------------------------------------------------------------

df = st.session_state.df_resultado_dim

if not df.empty:
    st.markdown("---")
    st.subheader("2. Consolidado y Analítica de Carga")

    cols_totales = st.columns(4)
    
    total_fob = df["Valor FOB (USD)"].sum()
    total_fletes = df["Sumatoria Fletes/Seguros/Otros (USD)"].sum()
    total_peso_bruto = df["Peso Bruto (Kgs)"].sum()
    total_peso_neto = df["Peso Neto (Kgs)"].sum()

    cols_totales[0].metric("Total Valor FOB (USD)", f"${total_fob:,.2f}")
    cols_totales[1].metric("Total Fletes/Seguros (USD)", f"${total_fletes:,.2f}")
    cols_totales[2].metric("Total Peso Bruto (Kgs)", f"{total_peso_bruto:,.2f}")
    cols_totales[3].metric("Total Peso Neto (Kgs)", f"{total_peso_neto:,.2f}")
    
    st.markdown("<br>", unsafe_allow_html=True)

    busqueda = st.text_input("🔍 Búsqueda rápida en el consolidado (Formulario, NIT, Importador, Manifiesto...)", "")

    df_vista = df.copy()
    if busqueda:
        mask = df_vista.apply(lambda fila: fila.astype(str).str.contains(busqueda, case=False, na=False).any(), axis=1)
        df_vista = df_vista[mask]

    st.dataframe(
        df_vista, 
        use_container_width=True, 
        height=420,
        column_config={
            "Valor FOB (USD)": st.column_config.NumberColumn(format="%.2f"),
            "Sumatoria Fletes/Seguros/Otros (USD)": st.column_config.NumberColumn(format="%.2f"),
            "Peso Bruto (Kgs)": st.column_config.NumberColumn(format="%.2f"),
            "Peso Neto (Kgs)": st.column_config.NumberColumn(format="%.2f"),
            "No. Bultos": st.column_config.NumberColumn(format="%d"),
        }
    )

    faltantes_totales = df[df["Campos_no_encontrados"] != ""]
    if not faltantes_totales.empty:
        with st.expander(f"⚠️ {len(faltantes_totales)} declaración(es) requieren revisión manual de campos"):
            st.dataframe(faltantes_totales[["Número de formulario", "Archivo", "Campos_no_encontrados"]], use_container_width=True)

    with st.expander("👁️ Inspección detallada por Formulario"):
        opciones = (df["Número de formulario"] + " — " + df["Archivo"]).tolist()
        seleccion = st.selectbox("Selecciona una declaración para ver su estructura completa", opciones)
        idx = opciones.index(seleccion)
        st.json(df.iloc[idx].to_dict())

    st.markdown("---")
    st.subheader("3. Exportación de Datos")
    df_export = df.drop(columns=["Campos_no_encontrados"])

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "⬇️ Descargar Reporte en Excel (.xlsx)",
            data=generar_excel(df_export),
            file_name=f"dim_zona_franca_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col2:
        st.download_button(
            "⬇️ Descargar Reporte en CSV (.csv)",
            data=generar_csv(df_export),
            file_name=f"dim_zona_franca_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
else:
    st.info("Carga tus documentos aduaneros para activar el panel de análisis y las métricas corporativas.")
